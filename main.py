from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, StreamingResponse
from google import genai
from google.genai import types
from openai import AsyncOpenAI
import httpx
from prompts import SYSTEM_PROMPT, build_evaluation_prompt
from dotenv import load_dotenv
import asyncio
import json
import os
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

load_dotenv(override=True)

app = FastAPI(title="AI Tools", version="0.2.0")

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

LLM_PROVIDER = os.environ.get("LLM_PROVIDER", "gemini").lower()

gemini_client: genai.Client | None = None
openai_client: AsyncOpenAI | None = None


def get_gemini_client() -> genai.Client:
    global gemini_client
    if gemini_client is None:
        api_key = os.environ.get("GEMINI_API_KEY")
        if not api_key:
            raise ValueError("GEMINI_API_KEY not set")
        gemini_client = genai.Client(api_key=api_key)
    return gemini_client


def get_openai_client() -> AsyncOpenAI:
    global openai_client
    if openai_client is None:
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OPENAI_API_KEY not set")
        # 顯式以 .env 的值控制組織：未設定時傳 None＝使用 key 的預設組織，
        # 可避免執行環境殘留的 OPENAI_ORG_ID 造成 invalid_organization (401)。
        openai_client = AsyncOpenAI(
            api_key=api_key,
            organization=os.environ.get("OPENAI_ORG_ID") or None,
        )
    return openai_client


async def web_search(query: str) -> str:
    """Execute a Google search via Serper.dev and return key information."""
    api_key = os.environ.get("SERPER_API_KEY")
    if not api_key:
        return "Search failed: SERPER_API_KEY not set"
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.post(
                "https://google.serper.dev/search",
                headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
                json={"q": query, "num": 5},
            )
            resp.raise_for_status()
            data = resp.json()

        parts = []

        # Answer Box — Google 精選摘要，最精華
        ab = data.get("answerBox", {})
        if ab:
            answer = ab.get("answer") or ab.get("snippet", "")
            if answer:
                link = ab.get("link", "")
                parts.append(f"[精選摘要] {answer} ({link})")

        # Knowledge Graph — 結構化資訊
        kg = data.get("knowledgeGraph", {})
        if kg:
            desc = kg.get("description", "")
            if desc:
                parts.append(f"[知識圖譜] {desc}")

        # Organic — 只取 snippet
        for r in data.get("organic", [])[:5]:
            snippet = r.get("snippet", "")
            link = r.get("link", "")
            if snippet:
                parts.append(f"- {snippet} ({link})")

        return "\n".join(parts) if parts else "No results found."
    except Exception as e:
        logger.warning("Web search failed for query %r: %s", query, e)
        return f"Search failed: {str(e)}"


web_search_declaration = types.FunctionDeclaration(
    name="web_search",
    description="Search the web using Google to find market data, competitor information, industry trends, and other real-time information to support the evaluation analysis.",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "query": types.Schema(
                type=types.Type.STRING,
                description="The search query string",
            ),
        },
        required=["query"],
    ),
)

search_tool = types.Tool(function_declarations=[web_search_declaration])

openai_tools = [
    {
        "type": "function",
        "function": {
            "name": "web_search",
            "description": "Search the web using Google to find market data, competitor information, industry trends, and other real-time information to support the evaluation analysis.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "The search query string",
                    }
                },
                "required": ["query"],
            },
        },
    }
]

MAX_SEARCH_ROUNDS = 3
MAX_TOTAL_SEARCHES = 5


async def extract_text_from_file(file: UploadFile) -> str:
    filename = file.filename.lower()
    content = await file.read()

    if filename.endswith(".txt") or filename.endswith(".md"):
        return content.decode("utf-8")

    if filename.endswith(".pdf"):
        from PyPDF2 import PdfReader
        import io

        reader = PdfReader(io.BytesIO(content))
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n".join(pages)

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                sheets.append(f"[工作表：{sheet_name}]\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(sheets) if sheets else "（Excel 檔案無內容）"

    raise ValueError(f"不支援的檔案格式：{filename}。僅支援 .txt、.md、.pdf、.xlsx")


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/api/health")
async def health():
    return {"status": "ok", "version": "0.2.0"}


def sse_event(event: str, data: dict) -> str:
    """Format a server-sent event."""
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"


async def with_keepalive(gen, interval=5):
    """Wrap an async generator with keep-alive heartbeats every `interval` seconds."""
    queue: asyncio.Queue[str | None] = asyncio.Queue()

    async def _produce():
        try:
            async for item in gen:
                await queue.put(item)
        finally:
            await queue.put(None)

    task = asyncio.create_task(_produce())
    while True:
        try:
            item = await asyncio.wait_for(queue.get(), timeout=interval)
        except asyncio.TimeoutError:
            yield ": keep-alive\n\n"
            continue
        if item is None:
            break
        yield item
    await task


@app.post("/api/evaluate")
async def evaluate(
    text: str = Form(default=""),
    file: UploadFile | None = File(default=None),
):
    parts = []

    if file and file.filename:
        try:
            file_text = await extract_text_from_file(file)
            parts.append(f"[檔案：{file.filename}]\n{file_text}")
        except ValueError as e:
            return StreamingResponse(
                iter([sse_event("error", {"error": str(e)})]),
                media_type="text/event-stream",
            )
        except Exception:
            return StreamingResponse(
                iter([sse_event("error", {"error": "檔案解析失敗，請確認檔案格式正確。"})]),
                media_type="text/event-stream",
            )

    if text.strip():
        parts.append(text.strip())

    if not parts:
        return StreamingResponse(
            iter([sse_event("error", {"error": "請輸入功能描述或上傳檔案。"})]),
            media_type="text/event-stream",
        )

    combined = "\n\n".join(parts)

    async def openai_stream():
        yield sse_event("status", {"message": "正在啟動分析..."})
        try:
            ai = get_openai_client()
        except ValueError:
            yield sse_event("error", {"error": "尚未設定 OpenAI API Key，請聯繫管理員。"})
            return

        try:
            yield sse_event("status", {"message": "正在呼叫 AI 模型..."})
            messages = [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": build_evaluation_prompt(combined)},
            ]

            total_searches = 0
            for round_i in range(MAX_SEARCH_ROUNDS + 1):
                use_tools = openai_tools if total_searches < MAX_TOTAL_SEARCHES else None
                response = await ai.chat.completions.create(
                    model="gpt-5.2",
                    messages=messages,
                    tools=use_tools,
                    temperature=0.0,
                )
                msg = response.choices[0].message
                messages.append(msg)

                if not msg.tool_calls:
                    break

                queries = []
                for tc in msg.tool_calls:
                    if total_searches >= MAX_TOTAL_SEARCHES:
                        messages.append(
                            {"role": "tool", "tool_call_id": tc.id, "content": "Search limit reached. Use existing information."}
                        )
                        continue
                    query = json.loads(tc.function.arguments)["query"]
                    queries.append((tc, query))
                    total_searches += 1
                    logger.info("Web search (%d/%d): %s", total_searches, MAX_TOTAL_SEARCHES, query)
                    yield sse_event("search", {"query": query})

                if queries:
                    results = await asyncio.gather(
                        *(web_search(q) for _, q in queries),
                        return_exceptions=True,
                    )
                    for (tc, _), result in zip(queries, results):
                        if isinstance(result, Exception):
                            result = f"Search failed: {result}"
                        messages.append(
                            {"role": "tool", "tool_call_id": tc.id, "content": result}
                        )
                yield sse_event("status", {"message": f"搜尋完成 ({total_searches}/{MAX_TOTAL_SEARCHES})，正在生成分析報告..."})

            # Stream the final result — use existing msg if available, otherwise stream
            if msg and msg.content:
                yield sse_event("result", {"result": msg.content})
            else:
                yield sse_event("status", {"message": "正在生成分析報告..."})
                stream = await ai.chat.completions.create(
                    model="gpt-5.2",
                    messages=messages,
                    temperature=0.0,
                    stream=True,
                )
                async for chunk in stream:
                    delta = chunk.choices[0].delta
                    if delta.content:
                        yield sse_event("delta", {"content": delta.content})
                yield sse_event("done", {})
        except Exception as e:
            yield sse_event("error", {"error": f"AI 分析時發生錯誤：{str(e)}"})

    async def gemini_stream():
        yield sse_event("status", {"message": "正在啟動分析..."})
        try:
            ai = get_gemini_client()
        except ValueError:
            yield sse_event("error", {"error": "尚未設定 Gemini API Key，請聯繫管理員。"})
            return

        try:
            yield sse_event("status", {"message": "正在呼叫 AI 模型..."})
            contents = [types.Content(role="user", parts=[types.Part.from_text(text=build_evaluation_prompt(combined))])]

            total_searches = 0
            for _ in range(MAX_SEARCH_ROUNDS + 1):
                use_tools = [search_tool] if total_searches < MAX_TOTAL_SEARCHES else None
                config_iter = types.GenerateContentConfig(
                    system_instruction=SYSTEM_PROMPT,
                    temperature=0.0,
                    tools=use_tools,
                )
                response = await ai.aio.models.generate_content(
                    model="gemini-3-flash-preview",
                    contents=contents,
                    config=config_iter,
                )

                function_calls = response.function_calls
                if not function_calls:
                    break

                contents.append(response.candidates[0].content)

                search_fcs = []
                for fc in function_calls:
                    if fc.name == "web_search":
                        if total_searches >= MAX_TOTAL_SEARCHES:
                            search_fcs_skip = types.Part.from_function_response(
                                name="web_search",
                                response={"result": "Search limit reached. Use existing information."},
                            )
                            contents.append(types.Content(role="user", parts=[search_fcs_skip]))
                            continue
                        query = fc.args.get("query", "")
                        search_fcs.append((fc, query))
                        total_searches += 1
                        logger.info("Web search (%d/%d): %s", total_searches, MAX_TOTAL_SEARCHES, query)
                        yield sse_event("search", {"query": query})

                if search_fcs:
                    search_results = await asyncio.gather(
                        *(web_search(q) for _, q in search_fcs),
                        return_exceptions=True,
                    )
                    function_responses = [
                        types.Part.from_function_response(
                            name="web_search",
                            response={"result": sr if not isinstance(sr, Exception) else f"Search failed: {sr}"},
                        )
                        for sr in search_results
                    ]
                    contents.append(types.Content(role="user", parts=function_responses))
                yield sse_event("status", {"message": f"搜尋完成 ({total_searches}/{MAX_TOTAL_SEARCHES})，正在分析中..."})

            yield sse_event("result", {"result": response.text})
        except Exception as e:
            yield sse_event("error", {"error": f"AI 分析時發生錯誤：{str(e)}"})

    generator = openai_stream if LLM_PROVIDER == "openai" else gemini_stream
    return StreamingResponse(
        with_keepalive(generator()),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


if __name__ == "__main__":
    import uvicorn

    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True, timeout_keep_alive=600)
